import io
import re
from datetime import datetime

import openpyxl
import requests
import streamlit as st
from bs4 import BeautifulSoup

EMA_REQUEST_TIMEOUT = 30
EMA_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def fetch_rmp_date(url: str) -> dict:
    """Fetch EMA page and extract the Risk Management Plan date.

    Returns a dict with keys: date_str, first_published, last_updated, error.
    """
    result = {
        "date_str": None,
        "first_published": None,
        "last_updated": None,
        "error": None,
    }

    if not url or not isinstance(url, str) or not url.strip():
        result["error"] = "Пустая ссылка"
        return result

    url = url.strip()

    try:
        resp = requests.get(url, headers=EMA_HEADERS, timeout=EMA_REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.RequestException as e:
        result["error"] = f"Ошибка загрузки: {e}"
        return result

    soup = BeautifulSoup(resp.text, "html.parser")
    text = soup.get_text(separator="\n")
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]

    rmp_idx = None
    for i, line in enumerate(lines):
        if re.search(r"EPAR\s*[-–—]\s*Risk[\s\-]+management[\s\-]+plan", line, re.IGNORECASE):
            rmp_idx = i
            break

    if rmp_idx is None:
        result["error"] = "RMP-секция не найдена на странице"
        return result

    date_re = re.compile(r"\d{2}/\d{2}/\d{4}")
    section_stop_re = re.compile(
        r"^(Product information|EPAR\s*[-–—]|All Authorised|Authorisation details"
        r"|Product details|Assessment history|More information)",
        re.IGNORECASE,
    )

    for line in lines[rmp_idx + 1 : rmp_idx + 15]:
        if "View" == line.strip():
            break
        if section_stop_re.search(line):
            break
        if line.startswith("First published"):
            m = date_re.search(line)
            if m and result["first_published"] is None:
                result["first_published"] = m.group(0)
        elif line.startswith("Last updated"):
            m = date_re.search(line)
            if m and result["last_updated"] is None:
                result["last_updated"] = m.group(0)
        else:
            m = date_re.search(line)
            if m:
                if result["first_published"] is None:
                    result["first_published"] = m.group(0)
                elif result["last_updated"] is None:
                    result["last_updated"] = m.group(0)

    if result["last_updated"]:
        result["date_str"] = result["last_updated"]
    elif result["first_published"]:
        result["date_str"] = result["first_published"]

    if result["date_str"] is None:
        result["error"] = "Даты RMP не найдены после заголовка секции"

    return result


def format_date_for_excel(date_str: str) -> str:
    """Convert dd/mm/yyyy to a more readable format."""
    try:
        dt = datetime.strptime(date_str, "%d/%m/%Y")
        return dt.strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return date_str or ""


def process_excel(file_bytes: bytes) -> tuple[bytes, list[dict]]:
    """Read uploaded Excel, fetch RMP dates, return filled Excel bytes and log."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header_row = 1
    first_data_row = 2

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]

    name_col = None
    ema_name_col = None
    rmp_col = None
    url_col = None

    for idx, h in enumerate(headers, start=1):
        if h is None:
            continue
        h_lower = str(h).lower().strip()
        if "наименование" in h_lower and "ema" not in h_lower and "fda" not in h_lower:
            if name_col is None:
                name_col = idx
        if "ema" in h_lower or "fda" in h_lower:
            ema_name_col = idx
        if "rmp" in h_lower or "дата версии" in h_lower:
            rmp_col = idx
        if "ссылк" in h_lower or "link" in h_lower or "url" in h_lower:
            url_col = idx

    if name_col is None:
        name_col = 1
    if ema_name_col is None:
        ema_name_col = 2
    if rmp_col is None:
        rmp_col = 3
    if url_col is None:
        url_col = 4

    log = []
    total = ws.max_row - first_data_row + 1
    progress_bar = st.progress(0, text="Обработка...")

    for row_idx in range(first_data_row, ws.max_row + 1):
        drug_name = ws.cell(row=row_idx, column=name_col).value or ""
        ema_name = ws.cell(row=row_idx, column=ema_name_col).value or ""
        url_val = ws.cell(row=row_idx, column=url_col).value

        display_name = str(ema_name or drug_name).strip()
        if not display_name:
            display_name = f"Строка {row_idx}"

        current = row_idx - first_data_row
        progress_bar.progress(
            current / total,
            text=f"Обработка {current + 1}/{total}: {display_name}",
        )

        if url_val and isinstance(url_val, str) and url_val.strip():
            link = url_val.strip()
        elif hasattr(ws.cell(row=row_idx, column=url_col), "hyperlink") and ws.cell(
            row=row_idx, column=url_col
        ).hyperlink:
            link = ws.cell(row=row_idx, column=url_col).hyperlink.target
        else:
            log.append(
                {"drug": display_name, "status": "Пропущено", "detail": "Нет ссылки"}
            )
            continue

        rmp_info = fetch_rmp_date(link)

        if rmp_info["date_str"]:
            formatted = format_date_for_excel(rmp_info["date_str"])
            ws.cell(row=row_idx, column=rmp_col).value = formatted
            source = "Last updated" if rmp_info["last_updated"] else "First published"
            log.append(
                {
                    "drug": display_name,
                    "status": "OK",
                    "detail": f"{source}: {formatted}",
                }
            )
        elif rmp_info["error"] and "Ошибка загрузки" in rmp_info["error"]:
            ws.cell(row=row_idx, column=rmp_col).value = "NA"
            log.append(
                {
                    "drug": display_name,
                    "status": "Ошибка",
                    "detail": rmp_info["error"],
                }
            )
        else:
            ws.cell(row=row_idx, column=rmp_col).value = "NA"
            log.append(
                {
                    "drug": display_name,
                    "status": "NA",
                    "detail": rmp_info["error"] or "RMP-секция не найдена",
                }
            )

    progress_bar.progress(1.0, text="Готово!")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), log


def main():
    st.set_page_config(
        page_title="EMA RMP Date Finder",
        page_icon="💊",
        layout="centered",
    )

    st.title("💊 EMA RMP Date Finder")
    st.markdown(
        "Загрузите Excel-файл со ссылками на страницы EMA — "
        "приложение автоматически найдёт даты **Risk Management Plan** "
        "для каждого препарата."
    )

    with st.expander("📋 Формат файла", expanded=False):
        st.markdown(
            """
            Excel-файл должен содержать **4 столбца**:
            1. **Наименование препарата** — название на русском
            2. **Наименование на сайте EMA/FDA** — название на английском
            3. **RMP (дата версии)** — сюда будет записана дата (заполняется автоматически)
            4. **Ссылка на страницу EMA** — URL страницы препарата на ema.europa.eu
            """
        )

    uploaded = st.file_uploader(
        "Выберите Excel-файл (.xlsx)",
        type=["xlsx"],
        help="Файл с 4 столбцами: название, название EMA, RMP дата, ссылка",
    )

    if uploaded is not None:
        st.success(f"Файл загружен: **{uploaded.name}** ({uploaded.size / 1024:.1f} КБ)")

        if st.button("🚀 Обработать", type="primary", use_container_width=True):
            with st.spinner("Загружаем данные с сайта EMA..."):
                file_bytes = uploaded.getvalue()
                result_bytes, log = process_excel(file_bytes)

            st.divider()
            st.subheader("📊 Результаты")

            ok_count = sum(1 for r in log if r["status"] == "OK")
            na_count = sum(1 for r in log if r["status"] == "NA")
            err_count = sum(1 for r in log if r["status"] == "Ошибка")
            skip_count = sum(1 for r in log if r["status"] == "Пропущено")

            cols = st.columns(4)
            cols[0].metric("Найдено", ok_count)
            cols[1].metric("NA (нет RMP)", na_count)
            cols[2].metric("Ошибки", err_count)
            cols[3].metric("Пропущено", skip_count)

            with st.expander("Подробный лог", expanded=True):
                for entry in log:
                    if entry["status"] == "OK":
                        st.markdown(f"✅ **{entry['drug']}** — {entry['detail']}")
                    elif entry["status"] == "NA":
                        st.markdown(
                            f"➖ **{entry['drug']}** — NA ({entry['detail']})"
                        )
                    elif entry["status"] == "Ошибка":
                        st.markdown(
                            f"❌ **{entry['drug']}** — {entry['detail']}"
                        )
                    else:
                        st.markdown(
                            f"⏭️ **{entry['drug']}** — {entry['detail']}"
                        )

            out_name = uploaded.name.rsplit(".", 1)[0] + "_RMP.xlsx"
            st.download_button(
                label="📥 Скачать заполненный Excel",
                data=result_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                type="primary",
                use_container_width=True,
            )

    st.divider()
    st.caption(
        "Данные извлекаются напрямую со страниц [ema.europa.eu](https://www.ema.europa.eu). "
        "Приложение ищет секцию «Risk management plan» и дату «Last updated»."
    )


if __name__ == "__main__":
    main()
